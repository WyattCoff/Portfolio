import sqlite3
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from Product import Product
from Transactions import Transaction


class SupermarketDAO:
    def __init__(self):
        self.database_name = 'supermarket.db'

    def add_product_to_db(self, product):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                # Use getter methods instead of direct attribute access
                cursor.execute('INSERT INTO Products (barcode, name, price) VALUES (?, ?, ?)',
                               (product.get_barcode(), product.get_description(), product.get_price()))
                conn.commit()
                return True
        except sqlite3.IntegrityError:
            print("A product with the same barcode already exists.")
            return False
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            return False

    def list_all_products(self):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT barcode, name, price FROM Products ORDER BY barcode ASC')
                rows = cursor.fetchall()

                # If no products are found in the database
                if len(rows) == 0:
                    print("No products found in the database.")
                    return []

                # Create Product instances from the database rows
                products = [Product(row[0], row[1], row[2]) for row in rows]
                return products
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            return []

    def find_product(self, barcode):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT barcode, name, price FROM Products WHERE barcode = ?', (barcode,))
                row = cursor.fetchone()
                if row:
                    return Product(row[0], row[1], row[2])
                else:
                    return None
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            return None

    def list_all_transactions(self):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT date, barcode, amount FROM Transactions ORDER BY date ASC')
                rows = cursor.fetchall()

                if not rows:
                    print("No transactions found in the database.")
                    return []

                transactions = [Transaction(row[0], row[1], row[2]) for row in rows]
                return transactions
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            return []

    def display_barchart_of_products_sold(self):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()

                # Aggregate the sum of product quantities sold
                cursor.execute('SELECT barcode, COUNT(barcode) FROM Transactions GROUP BY barcode')
                sales_data = cursor.fetchall()

                if not sales_data:
                    print("No sales data found.")
                    return

                # Populate dictionary with product barcodes and sales count
                sales_dict = {row[0]: row[1] for row in sales_data}

                # Fetch product names and prepare data for plotting
                product_names = []
                quantities_sold = []
                for barcode, count in sales_dict.items():
                    cursor.execute('SELECT name FROM Products WHERE barcode = ?', (barcode,))
                    product_name = cursor.fetchone()
                    if product_name:
                        product_names.append(product_name[0])
                        quantities_sold.append(count)

                # Create a workbook and select the active worksheet
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'supermarket_excel'

                # Add headers to the sheet
                sheet.append(['Product Name', 'Quantity Sold'])

                # Add product names and quantities sold to sheet
                for row in zip(product_names, quantities_sold):
                    sheet.append(row)

                # Create a bar chart
                chart = BarChart()
                chart.type = "col"  # Vertical bar chart
                chart.style = 10
                chart.title = "Bar Chart of Products Sold by Quantity"
                chart.y_axis.title = 'Quantity Sold'
                chart.x_axis.title = 'Product Name'

                # Reference for chart data, including the headers
                data = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row, max_col=2)
                # Make sure the categories (product names) also start from the second row to exclude headers
                cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
                chart.add_data(data, titles_from_data=False)  # Do not use the first row of data as titles
                chart.set_categories(cats)
                sheet.add_chart(chart, "E4")  # Place the chart starting at E4 to avoid overlap

                # Save the workbook to a file
                workbook.save('supermarket_sales_chart.xlsx')
                print("Excel chart created successfully.")

        except sqlite3.Error as e:
            print(f"An error occurred: {e}")

    def display_excel_report_of_transactions(self):
        transactions = self.list_all_transactions()
        if not transactions:
            print("No transactions to display in the report.")
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Transactions Report'
        sheet.append(['Date', 'Barcode', 'Amount'])

        for transaction in transactions:
            sheet.append([transaction.date, transaction.barcode, transaction.amount])

        workbook.save('Transactions_Report.xlsx')
        print("Transactions report created successfully.")

    def add_transaction_to_db(self, transaction):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                cursor.execute('INSERT INTO Transactions (date, barcode, amount) VALUES (?, ?, ?)',
                               (transaction.date, transaction.barcode, transaction.amount))
                conn.commit()
                print("Transaction saved successfully.")
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")

    def get_product_by_barcode(self, barcode):
        try:
            with sqlite3.connect(self.database_name) as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT barcode, name, price FROM Products WHERE barcode = ?', (barcode,))
                product_data = cursor.fetchone()
                return product_data
        except sqlite3.Error as e:
            print(f"An error occurred while fetching the product: {e}")
            return None
